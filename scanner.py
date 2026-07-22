import cv2
import pytesseract
import pandas as pd
import threading
import time
import tkinter as tk
from tkinter import filedialog, simpledialog
from difflib import SequenceMatcher
from openpyxl import Workbook
import os
import sys

# ✅ pegar caminho certo (EXE / PYTHON)
def base_path():
    return os.path.dirname(os.path.abspath(
        sys.executable if getattr(sys, 'frozen', False) else __file__
    ))

pasta_base = base_path()

# ✅ tesseract
pytesseract.pytesseract.tesseract_cmd = os.path.join(
    pasta_base, "tesseract", "tesseract.exe"
)

# ✅ tessdata (ESSENCIAL)
os.environ["TESSDATA_PREFIX"] = os.path.join(
    pasta_base, "tesseract", "tessdata"
)

# ------------------------------

lista = []
encontrados = []
duplicados = []
rodando = False

ultimo_nome = None
tempo_mostrar = 0
ultimo_processamento = 0


# 🔍 comparação
def parecido(a, b):
    return SequenceMatcher(None, a, b).ratio()


def buscar_nome(texto):
    texto = texto.upper()

    for nome in lista:
        if nome in texto or parecido(nome, texto) > 0.55:
            return nome

    return None


# 📁 carregar Excel
def carregar_excel():
    global lista
    caminho = filedialog.askopenfilename()

    if caminho:
        df = pd.read_excel(caminho, header=None)
        lista = df.iloc[:, 0].dropna().astype(str).str.upper().str.strip().tolist()

        status_label.config(text=f"📄 {len(lista)} nomes carregados")


# ✍️ digitar lista
def digitar_lista():
    global lista
    texto = simpledialog.askstring("Lista", "Digite nomes separados por vírgula")

    if texto:
        lista = [n.strip().upper() for n in texto.split(",")]
        status_label.config(text=f"📄 {len(lista)} nomes digitados")


# ▶️ iniciar
def iniciar():
    global rodando

    if not lista:
        status_label.config(text="⚠️ carregue lista primeiro")
        return

    rodando = True
    status_label.config(text="✅ rodando...")

    threading.Thread(target=scanner, daemon=True).start()


# ⛔ parar
def parar():
    global rodando
    rodando = False

    status_label.config(text="💾 salvando...")
    time.sleep(2)

    salvar_excel()
    status_label.config(text="✅ Excel salvo com sucesso")


# 📷 abrir camera
def abrir_camera():
    for i in range(5):
        cap = cv2.VideoCapture(i)
        time.sleep(0.5)
        if cap.isOpened():
            return cap
    return None


# 🎥 scanner
def scanner():
    global rodando, ultimo_nome, tempo_mostrar, ultimo_processamento

    cap = abrir_camera()

    if cap is None:
        status_label.config(text="❌ câmera não abriu")
        rodando = False
        return

    falhas = 0

    while rodando:
        ret, frame = cap.read()

        if not ret:
            falhas += 1
            if falhas < 15:
                time.sleep(0.2)
                continue
            else:
                break
        else:
            falhas = 0

        tempo = time.time()

        if tempo - ultimo_processamento > 2:
            ultimo_processamento = tempo

            gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
            gray = cv2.resize(gray, None, fx=1.5, fy=1.5)

            config = r'--oem 3 --psm 6 -l eng'

            texto = pytesseract.image_to_string(gray, config=config).upper()

            nome = buscar_nome(texto)

            if nome:
                ultimo_nome = nome
                tempo_mostrar = tempo

                if nome not in encontrados:
                    encontrados.append(nome)
                else:
                    if nome not in duplicados:
                        duplicados.append(nome)

                atualizar_status()

        # ✅ overlay
        if ultimo_nome and (tempo - tempo_mostrar < 3):

            cv2.putText(frame, "ENCONTRADO", (30, 50),
                        cv2.FONT_HERSHEY_SIMPLEX, 1, (0, 255, 0), 3)

            cv2.putText(frame, f"{ultimo_nome}", (30, 100),
                        cv2.FONT_HERSHEY_SIMPLEX, 0.7, (0, 255, 0), 2)

            cv2.putText(frame, f"Total: {len(lista)}", (30, 150),
                        cv2.FONT_HERSHEY_SIMPLEX, 0.6, (255, 255, 255), 2)

            cv2.putText(frame, f"OK: {len(encontrados)}", (30, 180),
                        cv2.FONT_HERSHEY_SIMPLEX, 0.6, (0, 255, 0), 2)

            cv2.putText(frame, f"Duplicados: {len(duplicados)}", (30, 210),
                        cv2.FONT_HERSHEY_SIMPLEX, 0.6, (0, 255, 255), 2)

        else:
            cv2.putText(frame, "PROCURANDO...", (30, 50),
                        cv2.FONT_HERSHEY_SIMPLEX, 1, (0, 0, 255), 3)

        cv2.imshow("Scanner", frame)

        if cv2.waitKey(1) & 0xFF == 27:
            break

    cap.release()
    cv2.destroyAllWindows()
    rodando = False


# 📊 status
def atualizar_status():
    faltando = len([n for n in lista if n not in encontrados])

    status_label.config(
        text=f"📄 {len(lista)} | ✅ {len(encontrados)} | ⚠️ {len(duplicados)} | ❌ {faltando}"
    )


# 💾 EXCEL (100% SEM CORROMPER)
def salvar_excel():
    faltando = [n for n in lista if n not in encontrados]

    try:
        wb = Workbook()

        # remove aba padrão
        wb.remove(wb.active)

        aba1 = wb.create_sheet("Encontrados")
        aba2 = wb.create_sheet("Duplicados")
        aba3 = wb.create_sheet("Faltantes")

        aba1.append(["Encontrados"])
        for n in encontrados:
            aba1.append([n])

        aba2.append(["Duplicados"])
        for n in duplicados:
            aba2.append([n])

        aba3.append(["Faltantes"])
        for n in faltando:
            aba3.append([n])

        wb.save("resultado.xlsx")
        wb.close()

        print("✅ Excel salvo perfeito")

    except Exception as e:
        print("Erro Excel:", e)


# 🖥️ interface
app = tk.Tk()
app.title("Scanner de Certificados")
app.geometry("350x300")

tk.Label(app, text="Scanner de Certificados", font=("Arial", 14)).pack(pady=10)

tk.Button(app, text="Carregar Excel", command=carregar_excel).pack(pady=5)
tk.Button(app, text="Digitar Lista", command=digitar_lista).pack(pady=5)

tk.Button(app, text="Iniciar", command=iniciar, bg="green").pack(pady=5)
tk.Button(app, text="Parar", command=parar, bg="red").pack(pady=5)

status_label = tk.Label(app, text="Nenhuma lista carregada")
status_label.pack(pady=10)

app.mainloop()